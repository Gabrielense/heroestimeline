# -*- coding: utf-8 -*-
"""Re-read the source Google Sheet and rewrite the data block inside index.html.

    py build/sync.py            # download a fresh copy, then inject
    py build/sync.py --offline  # reuse build/sheet.xlsx

Everything outside the <script id="timeline-data"> block is left untouched, so
the page itself can be hand-edited freely.

Requires: openpyxl  (pip install openpyxl)
"""
import argparse, datetime, json, os, re, sys, urllib.request

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import season4

SHEET_ID = "1Ci6zyz2nhjgrCurrhSDrD_jvtfXBLq_l9CWLpKDe3NE"
XLSX_URL = "https://docs.google.com/spreadsheets/d/%s/export?format=xlsx" % SHEET_ID

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "sheet.xlsx")
HTML = os.path.join(HERE, os.pardir, "index.html")

# cell fill -> era id. The sheet's legend row uses one green ramp for the five
# volumes, gold for Heroes Reborn, and a near-black for the off-air stretches.
ERAS = {
    "FFCCFFCC": "v1", "FF99FF99": "v2", "FF66FF66": "v3",
    "FF33CC33": "v4", "FF008000": "v5",
    "FFFFE599": "hr", "FFFFE699": "hr",
    "FFCCCCCC": "gap", "FF262626": "gap",
}

# The show went off the air, but the releases did not, and a timeline of
# releases has no empty stretch to name. Volume 5 runs to the last thing it put
# out -- the complete-series set of 16 November 2010 -- and Heroes Reborn starts
# with the first thing it put out, the teaser NBC slipped into the Olympics on
# 23 February 2014. So the near-black rows are not an era of their own: each one
# belongs to whichever side of that line it falls on.
VOLUME_5_ENDS = "2010-11-16"
REBORN_BEGINS = "2014-02-23"


def resolve_gaps(entries):
    """Give every off-air row the era on its side of the Reborn announcement."""
    n, prev = 0, "v1"
    for e in entries:
        if e["e"] == "gap":
            # the undated markers have no side of the line to fall on, so they
            # stay with the era they interrupt
            e["e"] = ("hr" if e["d"] >= REBORN_BEGINS else "v5") if e.get("d") else prev
            n += 1
        prev = e["e"]
    return n
# --- Internet Archive -------------------------------------------------------
# Entries that can be played in place. The page turns these into an
# archive.org embed; nothing is rehosted here, it streams from archive.org.
#
# The archive item numbers Heroes Unmasked's first season differently from the
# sheet - it counts "A Heroes Welcome" and "The Story So Far" as episodes of
# their own - so these are matched on title, not on episode code. Regenerate
# with build/link_archive.py if the item ever changes.
UNMASKED_ID = "heroes.unmasked.behind-the-scenes"

UNMASKED_FILES = {
    "A Bug's Life":                    "Season 3/Heroes.Unmasked.S03E07.A.Bugs.Life.WS.PDTV.XviD-BiA.mp4",
    "A Heroes Welcome":                "Season 1/Heroes.Unmasked.S01E09.A.Heroes.Welcome.WS.PDTV.XviD-BiA.mp4",
    "A New Dawn (UK premiere date 7/25/2007)": "Season 1/Heroes.Unmasked.S01E01.A.New.Dawn.HDTV.XviD-PVR.mp4",
    "Action!":                         "Season 3/Heroes.Unmasked.S03E09.Action.WS.PDTV.XviD-BiA.mp4",
    "Bad Company":                     "Season 1/Heroes.Unmasked.S01E17.Bad.Company.HDTV.XviD-BiA.mp4",
    "Dark Angel Gabriel":              "Season 1/Heroes.Unmasked.S01E10.Dark.Angel.Gabriel.HDTV.XviD-BiA.mp4",
    "Double Trouble":                  "Season 1/Heroes.Unmasked.S01E06.Double.Trouble.HDTV.XviD-BiA.mp4",
    "Finale":                          "Season 1/Heroes.Unmasked.S01E21.Finale.WS.PDTV.XviD-BiA.mp4",
    "From Heroes to Villains":         "Season 2/Heroes.Unmasked.S02E10.From.Heroes.To.Villains.WS.PDTV.XviD-BiA.mp4",
    "Generations":                     "Season 2/Heroes.Unmasked.S02E02.Generations.WS.PDTV.XviD-BiA.mp4",
    "Growing Pains":                   "Season 1/Heroes.Unmasked.S01E02.Growing.Pains.HDTV.XviD-BiA.mp4",
    "Head to Head":                    "Season 1/Heroes.Unmasked.S01E14.Head.to.Head.HDTV.XviD-BiA.mp4",
    "Heroes Return":                   "Season 2/Heroes.Unmasked.S02E00.Heroes.Return.WS.PDTV.XviD-BiA.mp4",
    "Heroes by Design":                "Season 3/Heroes.Unmasked.S03E11.Heroes.By.Design.WS.PDTV.XviD-PVR.mp4",
    "Heroes on the Run":               "Season 3/Heroes.Unmasked.S03E13.Heroes.On.The.Run.WS.PDTV.XviD-BiA.mp4",
    "Hiro Worship":                    "Season 1/Heroes.Unmasked.S01E04.Hiro.Worship.HDTV.XviD-BiA.avi.mp4",
    "Isaac":                           "Season 1/Heroes.Unmasked.S01E18.Isaac.HDTV.XviD-BiA.mp4",
    "Japanese Idol":                   "Season 2/Heroes.Unmasked.S02E01.Japanese.Idol.WS.PDTV.XviD-BiA.mp4",
    "Let's Get Physical":              "Season 3/Heroes.Unmasked.S03E10.Let's.Get.Physical.WS.PDTV.XviD-BiA.mp4",
    "Missing Links":                   "Season 3/Heroes.Unmasked.S03E08.Missing.Links.WS.PDTV.XviD-BiA.mp4",
    "Mohinder's Journey":              "Season 1/Heroes.Unmasked.S01E11.Mohinders.Journey.WS.PDTV.XviD-BiA.mp4",
    "New Heroes On The Block":         "Season 3/Heroes.Unmasked.S03E02.New.Heroes.On.The.Block.WS.PDTV.XviD-BiA.mp4",
    "New World Disorder":              "Season 2/Heroes.Unmasked.S02E07.New.World.Disorder.WS.PDTV.XviD-BiA.mp4",
    "On a Heroic Scale":               "Season 3/Heroes.Unmasked.S03E03.On.A.Heroic.Scale.WS.PDTV.XVID-BiA.mp4",
    "Opening Pandora's Box":           "Season 3/Heroes.Unmasked.S03E01.Opening.Pandoras.Box.WS.PDTV.XviD-BiA.mp4",
    "Painting the Future":             "Season 1/Heroes.Unmasked.S01E05.Painting.The.Future.WS.PDTV.XviD-UKN.mp4",
    "Playing God":                     "Season 3/Heroes.Unmasked.S03E04.Playing.God.WS.PDTV.XVID-BIA.mp4",
    "Sets and the City":               "Season 1/Heroes.Unmasked.S01E15.Sets.And.The.City.HDTV.XviD-BiA.mp4",
    "Shock of the Old":                "Season 3/Heroes.Unmasked.S03E05.Shock.Of.The.Old.WS.PDTV.XviD-BiA.mp4",
    "Sweet Dreams":                    "Season 2/Heroes.Unmasked.S02E05.Sweet.Dreams.WS.PDTV.XviD-BiA.mp4",
    "Teenage Kicks":                   "Season 3/Heroes.Unmasked.S03E06.Teenage.Kicks.WS.PDTV.XviD-BiA.mp4",
    "Telling Tales":                   "Season 1/Heroes.Unmasked.S01E13.Telling.Tales.HDTV.XviD-BiA.mp4",
    "The Casting Couch":               "Season 2/Heroes.Unmasked.S02E08.The.Casting.Couch.WS.PDTV.XviD-BiA.mp4",
    "The Director's Cut":              "Season 1/Heroes.Unmasked.S01E20.The.Director's.Cut.HDTV.XviD-BiA.mp4",
    "The Dreamer":                     "Season 1/Heroes.Unmasked.S01E03.The.Dreamer.HDTV.XviD-BiA.mp4",
    "The H.R.G. File":                 "Season 1/Heroes.Unmasked.S01E08.The.H.R.G.File.WS.PDTV.XviD-BiA.mp4",
    "The Invisible Touch":             "Season 1/Heroes.Unmasked.S01E12.The.Invisible.Touch.HDTV.XviD-BiA.mp4",
    "The Music of Heroes":             "Season 3/Heroes.Unmasked.S03E12.The.Music.Of.Heroes.WS.PDTV.XviD-BiA.mp4",
    "The Story So Far":                "Season 1/Heroes.Unmasked.S01E22.The.Story.So.Far.WS.PDTV.XviD-NOsegmenT.mp4",
    "Through Heroes Eyes":             "Season 2/Heroes.Unmasked.S02E09.Through.Heroes.Eyes.WS.PDTV.XviD-BiA.mp4",
    "Tomorrow's World":                "Season 1/Heroes.Unmasked.S01E19.Tomorrows.World.WS.PDTV.XviD-BiA.mp4",
    "Travelling in Style":             "Season 2/Heroes.Unmasked.S02E06.Travelling.In.Style.WS.PDTV.XviD-BiA.mp4",
    "Turning Tides":                   "Season 1/Heroes.Unmasked.S01E16.Turning.Tides.HDTV.XviD-BiA.mp4",
    "Voices":                          "Season 1/Heroes.Unmasked.S01E07.Voices.HDTV.XviD-BiA.mp4",
    "What the Butlers Saw":            "Season 2/Heroes.Unmasked.S02E03.What.The.Butlers.Saw.WS.PDTV.XviD-BiA.mp4",
    "When Worlds Collide":             "Season 2/Heroes.Unmasked.S02E04.When.Worlds.Collide.WS.PDTV.XviD-BiA.mp4",
}

# Things that live in an archive item of their own, which the sheet does not
# link. A single-video item like these plays in the panel through the embed;
# see the note above videoFor() in index.html for why that differs from the
# Unmasked episodes. Inside the Eclipse #09 is not on archive.org at all.
EXTRA_LINKS = {
    (194, "bts", "Inside the Eclipse #02"):
        "https://archive.org/details/Heroes_Reborn_-_Inside_the_Eclipse_Episode_2_-_Odessa",
    # The pilot NBC never aired -- longer, differently cut, and the only episode
    # of the series that was never sold anywhere.
    (4, "ep", "Unaired Pilot"):
        "https://archive.org/details/heroes-original-unaired-pilot-english-subtitles",
}


def attach_archive(row, key, items, applied):
    """Give an item an `ia` embed path when we know where it lives."""
    for it in items:
        url = EXTRA_LINKS.get((row, key, it["t"]))
        if url:
            applied.add(it["t"])
            it.setdefault("l", url)
        if key == "bts" and it["t"] in UNMASKED_FILES:
            it["ia"] = UNMASKED_ID + "/" + UNMASKED_FILES[it["t"]]
    return items


# The sixth column used to be Heroes Unmasked and nothing else, so the sheet
# gives its episodes bare titles. It is now the whole behind-the-scenes column
# -- Inside Heroes, the disc featurettes, Countdown to the Premiere -- and
# "Finale" or "Isaac" on their own no longer say what they are. Naming them
# here rather than in the sheet keeps the archive map above matching on the
# title the archive item uses.
UNM_CODE = re.compile(r"^\dx\d\d$")


def name_unmasked(key, items):
    if key != "bts":
        return items
    for it in items:
        if it["t"] in UNMASKED_FILES or UNM_CODE.match(it.get("c") or ""):
            it["t"] = "Unmasked: " + it["t"]
    return items


KEYS = ["date", "ep", "gn", "web", "evo", "bts", "phys", "misc"]
HEADER_ROW = 3          # column titles (and their source links) live here
FIRST_DATA_ROW = 4

# --- corrections -------------------------------------------------------------
# The sheet carries a handful of misspellings and two factual slips. Rather than
# edit the sheet (and lose them on the next sync), they are patched here on the
# way through. Every entry was checked against Wikipedia's episode and graphic
# novel lists. Delete a line to let the sheet's own wording through again.

# keyed by (column, exact text in the sheet) -> corrected text
TITLE_FIXES = {
    # episodes
    ("ep", "Better Halves."):                   "Better Halves",
    ("ep", "Seven Minutes to Midnightc"):       "Seven Minutes to Midnight",
    ("ep", "Landside"):                         "Landslide",
    ("ep", "How to Stop na Explosing Man"):     "How to Stop an Exploding Man",
    ("ep", "The Kidness of Stranger"):          "The Kindness of Strangers",
    ("ep", "Truths & Consequences"):            "Truth & Consequences",
    ("ep", "Jump, Push Fall"):                  "Jump, Push, Fall",
    # graphic novels
    ("gn", "Isaac's Fist Time"):                        "Isaac's First Time",
    ("gn", "How Do You Stop na Explosing Man?, Part 1"): "How Do You Stop an Exploding Man?, Part 1",
    ("gn", "How Do You Stop na Explosing Man?, Part 2"): "How Do You Stop an Exploding Man?, Part 2",
    ("gn", "The Death of Hana Gilteman, Part 1"):       "The Death of Hana Gitelman, Part 1",
    ("gn", "The Death of Hana Gilteman, Part 2"):       "The Death of Hana Gitelman, Part 2",
    ("gn", "Flying Bird"):                              "Flying Blind",
    ("gn", "Amost Famous"):                             "Almost Famous",
    ("gn", "The Natural Order of The Things"):          "The Natural Order of Things",
    # #166-173 are "From the Files of Primatech", not "of Company"
    ("gn", "From the Files of Company, Part 1"): "From the Files of Primatech, Part 1",
    ("gn", "From the Files of Company, Part 2"): "From the Files of Primatech, Part 2",
    ("gn", "From the Files of Company, Part 3"): "From the Files of Primatech, Part 3",
    ("gn", "From the Files of Company, Part 4"): "From the Files of Primatech, Part 4",
    ("gn", "From the Files of Company, Part 5"): "From the Files of Primatech, Part 5",
    ("gn", "From the Files of Company, Part 6"): "From the Files of Primatech, Part 6",
    ("gn", "From the Files of Company, Part 7"): "From the Files of Primatech, Part 7",
    ("gn", "From the Files of Company, Part 8"): "From the Files of Primatech, Part 8",
    # evolutions / iStory
    ("evo", "brianundaunted.imeem.com."):               "brianundaunted.imeem.com",
    ("evo", "Primehearst & Primatech SMS and emails"):  "Pinehearst & Primatech SMS and emails",
    ("evo", "Operation Splinte/Operation Bad Blood"):   "Operation Splinter/Operation Bad Blood",
    # misc
    ("misc", "Heroes Reborn teaser at 2014 Winter Olympics Comercials"):
             "Heroes Reborn teaser in 2014 Winter Olympics commercials",
    # Named for the site it ran on and for the files themselves, since it is
    # remembered under both and searched for under both.
    ("evo", "Global News Interactive's Richard Drucker reports (The Drucker Files)"):
            "Global News Interactive / The Drucker Files",
}

# keyed by (row, column, exact text) -> corrected text, for cases where the same
# wrong string is correct somewhere else in the sheet
ROW_FIXES = {
    # The Official Magazine ran twelve bi-monthly issues to December 2009; #11 is
    # listed twice, and the October 2009 one is #12.
    (159, "phys", "Heroes: The Official Magazine #11"): "Heroes: The Official Magazine #12",
    # "Starting Over" is right for #159; #162 is "Second Chances"
    (172, "gn", "Starting Over"): "Second Chances",
}

# --- columns the sheet files under the wrong medium --------------------------
# The sheet has a single "misc" column for everything that is not an episode, a
# novel, a webisode, an ARG chapter or a disc, and it had grown into three
# unrelated things at once. They are split here rather than in the sheet:
#
#   making-of material  -> behind the scenes, beside Heroes Unmasked
#   in-fiction artefacts -> Evolutions, where the rest of the ARG already is
#   merchandise          -> the physical releases, which have somewhere to buy
#
# What stays in Misc. is the real-world material: broadcasts *about* the show,
# the tour, the adverts, the fan club, the announcements.
#
#   (column, exact text in the sheet) -> {k: new column, t: new title,
#                                         c: code to give it}
RECLASS = {
    ("misc", "9thwonders.com"):                   {"k": "evo"},
    ("misc", "Claire & the Cat flash animation"): {"k": "evo"},
    ("misc", "Heroes Reborn app"):                {"k": "evo"},
    ("misc", "Inside Heroes (Heroes Webisode Behind the Story)"):
        {"k": "bts", "t": "Inside Heroes: Heroes Webisode Behind the Story"},
    # numbered like the Unmasked episodes it now sits beside
    ("misc", "Heroes: Countdown to the Premiere (#3x00)"):
        {"k": "bts", "t": "Heroes: Countdown to the Premiere", "c": "3x00"},
    ("misc", "Topps Heroes Series 1 Trading Card Set"): {"k": "phys"},
    ("misc", "Topps Heroes Series 2 Trading Card Set"): {"k": "phys"},
    ("misc", "Action figures: Series 1"):               {"k": "phys"},
    ("misc", "Action figures: Series 2"):               {"k": "phys"},
    # a 566-page TV Guide book, not a broadcast -- it belongs with the other
    # books rather than with the adverts
    ("misc", "Heroes: Some People Are Born to Be Extraordinary"): {"k": "phys"},
}

# Rows the sheet packs into one line and that are listed out in
# build/data/extra_releases.json instead.
DROP = {
    ("misc", "Inside Heroes #1 to #8"),
}


def apply_reclass(entries):
    moved, dropped = 0, 0
    for entry in entries:
        cols = entry.get("c") or {}
        landing = {}
        for key, items in list(cols.items()):
            keep = []
            for it in items:
                if (key, it["t"]) in DROP:
                    dropped += 1
                    continue
                rule = RECLASS.get((key, it["t"]))
                if not rule:
                    keep.append(it)
                    continue
                if rule.get("t"):
                    it["t"] = rule["t"]
                if rule.get("c"):
                    it["c"] = rule["c"]
                landing.setdefault(rule["k"], []).append(it)
                moved += 1
            if keep:
                cols[key] = keep
            else:
                del cols[key]
        for key, items in landing.items():
            cols.setdefault(key, []).extend(items)
    return moved, dropped


# Undated rows in the date column. DATE_LABELS ones read as the row's own date;
# anything else becomes a full-width divider above the row.
DATE_LABELS = {
    "BEFORE PREMIERE": "Pre-pilot",
}

# --- additions ---------------------------------------------------------------
# Releases the sheet has no row for. Each one names the week it belongs to and
# the column it goes in; a week that does not exist yet is created and takes its
# era from the week before it. Loaded from build/data/additions.json so a run of
# a collector can add to them without anyone editing this file.
#
# The shape of one addition:
#   {"d": "2010-03-10", "k": "gn", "c": "167b", "t": "Novel Approach: ...",
#    "l": "https://..."}


# --- rows that pack several releases into one -------------------------------
# The sheet writes a week's worth of parts as a single cell when they went out
# together -- "Dark Matters, Parts 2, 3, 4, 5 & 6". They are still five
# webisodes with five titles and five summaries, so they become five entries.
# Listed out by hand rather than parsed: there are five of them, and guessing
# at "Parts 1 & 2" versus "Operation Splinter/Operation Bad Blood" would be
# more code and less certain than writing them down.
SPLITS = {
    ("web", "10-11"): [("10", "Hard Knox, Part 1"),
                       ("11", "Hard Knox, Part 2")],
    ("web", "13-14"): [("13", "Hard Knox, Part 3"),
                       ("14", "Hard Knox, Part 4")],
    ("web", "32-36"): [("32", "Dark Matters, Part 2"),
                       ("33", "Dark Matters, Part 3"),
                       ("34", "Dark Matters, Part 4"),
                       ("35", "Dark Matters, Part 5"),
                       ("36", "Dark Matters, Part 6")],
    ("web", "37-38"): [("37", "Damen Peak, Part 1"),
                       ("38", "Damen Peak, Part 2")],
    ("evo", "Cap. 106/206"): [("Cap. 106", "Operation Splinter"),
                              ("Cap. 206", "Operation Bad Blood")],
}


def apply_splits(entries):
    made = 0
    for entry in entries:
        for key, items in list((entry.get("c") or {}).items()):
            out = []
            for item in items:
                parts = SPLITS.get((key, item.get("c") or ""))
                if not parts:
                    out.append(item)
                    continue
                for code, title in parts:
                    fresh = dict(item)
                    fresh["c"], fresh["t"] = code, title
                    out.append(fresh)
                made += len(parts) - 1
            entry["c"][key] = out
    return made


# --- the ones that were never numbered --------------------------------------
# Three novels sit outside the run: two the sheet gives no issue number at all,
# and the Hiro reprint collection, which only ever had a number because it had
# to go somewhere. They read as "Bonus" rather than as a number or a blank.
# The page looks their blurbs up by title, since "Bonus" is not a key.
BONUS = [
    ("gn", "The Rogue"),
    ("gn", "The Last Shangri-La"),
    ("gn", "Novel Approach: The Hiro Collection"),
]


def apply_bonus(entries):
    marked = 0
    for entry in entries:
        for key, title in BONUS:
            for item in (entry.get("c") or {}).get(key, []):
                if item.get("t") == title:
                    item["c"] = "Bonus"
                    marked += 1
    return marked


# --- weeks the sheet files wrong -------------------------------------------
# Only where two independent sources agree against it. Both of these were found
# by build/gn_dates.py against Wikipedia's list and confirmed against
# User:Iheartheroes' release dates; both are given as the Monday of the right
# week, since that is how every other row in the sheet is filed.
#
# Everything else the two lists disagree about is a day either side of a Monday
# -- the novels went out on Tuesdays -- which is a convention, not an error.
DATE_MOVES = {
    ("gn", "Viewpoints"): "2008-11-03",
    ("gn", "From the Files of Primatech, Part 8"): "2010-06-07",
    # Heroes Wiki dates the Drucker report to the day Hana Gitelman posted the
    # four clips to her blog, 13 December 2007 -- the Thursday of the following
    # week from where the sheet files it.
    ("evo", "Global News Interactive / The Drucker Files"): "2007-12-10",
    # Three more the wiki dates differently, and it is the better source: it was
    # written while these were running. Each entry's panel links the page the
    # date comes from.
    #   Are You a Hero?  the quiz's own Evolutions box gives 8 January 2007,
    #                    mid-season-one hiatus, not the December before it
    #   Habbo            the campaign the wiki records runs 23 January to
    #                    9 February 2009, its box dated to the 26th -- nowhere
    #                    near the Volume Three premiere week the sheet files it in
    #   Gemini           out on Xbox One, Steam and PC on 19 January 2016, with
    #                    the PS4 version a week later
    ("evo", "Are You A Hero?"): "2007-01-08",
    ("evo", "Habbo's Interactivity"): "2009-01-26",
    ("evo", "Gemini: Heroes Reborn (PC/PS4/XBOX One)"): "2016-01-19",
}


def apply_moves(entries):
    """Lift an item out of the week the sheet gives it and into the right one."""
    moved = 0
    for (key, title), target in DATE_MOVES.items():
        found = None
        for entry in entries:
            items = (entry.get("c") or {}).get(key)
            if not items:
                continue
            for item in items:
                if item.get("t") == title:
                    found = (entry, item)
                    break
            if found:
                break
        if not found:
            print("  note: nothing to move for %r" % (title,), file=sys.stderr)
            continue
        entry, item = found
        if entry.get("d") == target:
            continue
        entry["c"][key].remove(item)
        if not entry["c"][key]:
            del entry["c"][key]
        _place(entries, target, key, item)
        moved += 1
    return moved


def _place(entries, date, key, item):
    """Put an item in the week for `date`, making that week if it is not there."""
    for entry in entries:
        if entry.get("d") == date:
            entry.setdefault("c", {}).setdefault(key, []).append(item)
            return
    at, era = len(entries), "gap"
    for i, entry in enumerate(entries):
        if entry.get("d") and entry["d"] < date:
            at, era = i + 1, entry["e"]
    entries.insert(at, {"r": 2000 + at, "e": era, "d": date,
                        "c": {key: [item]}})


def load_additions():
    path = os.path.join(HERE, "data", "additions.json")
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def apply_additions(entries, additions):
    """Fold hand-written and collected extras into the weeks they belong to."""
    by_date = {e["d"]: e for e in entries if e.get("d")}
    added = 0
    for item in additions:
        date, key = item.get("d"), item.get("k")
        if not date or not key:
            continue
        row = dict(item)
        row.pop("d", None)
        row.pop("k", None)
        entry = by_date.get(date)
        if entry is None:
            # A week of its own, slotted in by hand rather than by sorting: the
            # undated rows -- Pre-pilot, CANCELLATION, HEROES REBORN -- carry no
            # date to sort on and would be flung to the top of the page.
            at, era = len(entries), "gap"
            for i, e in enumerate(entries):
                if e.get("d") and e["d"] < date:
                    at, era = i + 1, e["e"]
            entry = {"r": 1000 + added, "e": era, "d": date, "c": {}}
            entries.insert(at, entry)
            by_date[date] = entry
        entry.setdefault("c", {}).setdefault(key, []).append(row)
        added += 1
    return added

CODE_RE = re.compile(
    r"^(#?(?:HR)?\d+x\d+|#\d+(?:-#?\d+)?|Cap\.\s*\d+(?:/\d+)?|\d+x\d+)\s*[:–-]\s*(.+)$")


def norm_date(raw):
    """The sheet's locale is day-first but the author typed US month-first dates,
    so every value Sheets actually parsed came out with day and month swapped.
    Swap them back. Values whose first number is > 12 were never parsed and are
    still literal M/D/Y strings. Sanity check: with this rule every date in the
    sheet lands in strict chronological order."""
    if raw is None:
        return None
    if isinstance(raw, datetime.datetime):
        return datetime.date(raw.year, raw.day, raw.month)
    if isinstance(raw, datetime.date):
        return datetime.date(raw.year, raw.day, raw.month)
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", str(raw).strip())
    if m:
        return datetime.date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
    return None


def split_items(text):
    """One cell can hold several releases joined by ' + '. Pull the leading
    episode/issue/chapter code out of each so the page can style it."""
    out = []
    parts = [p.strip() for p in re.split(r"\s+\+\s+", text) if p.strip()]
    for i, part in enumerate(parts):
        # "Inside the Eclipse #01 + #02" - the bare number inherits the series name
        if i and re.match(r"^#\d+$", part):
            prev = re.match(r"^(.*?)#\d+\s*$", parts[i - 1])
            if prev and prev.group(1).strip():
                part = prev.group(1).strip() + " " + part
        m = CODE_RE.match(part)
        if m:
            out.append({"c": m.group(1).lstrip("#").strip(), "t": m.group(2).strip()})
        else:
            out.append({"t": part})
    return out


def apply_fixes(row, key, items, applied):
    for it in items:
        for table, k in ((ROW_FIXES, (row, key, it["t"])), (TITLE_FIXES, (key, it["t"]))):
            if k in table:
                applied.add(k[-1])
                it["t"] = table[k]
                break
    return items


def cell_fill(cell):
    f = cell.fill
    if f and f.fgColor and f.fgColor.type == "rgb" and f.fgColor.rgb not in (None, "00000000"):
        return f.fgColor.rgb
    return None


def build(path):
    import openpyxl
    ws = openpyxl.load_workbook(path).active

    header = [ws.cell(row=HEADER_ROW, column=i + 1) for i in range(8)]
    col_links = {KEYS[i]: header[i].hyperlink.target
                 for i in range(8) if header[i].hyperlink}

    entries, prev, applied = [], None, set()
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        cells = [ws.cell(row=r, column=i + 1) for i in range(8)]
        if all(c.value is None for c in cells):
            continue
        fill = next((cell_fill(c) for c in cells if cell_fill(c)), None)
        entry = {"r": r, "e": ERAS.get(fill, "gap")}

        dt = norm_date(cells[0].value)
        if dt:
            if prev and dt < prev:
                print("  warning: row %d (%s) is out of order" % (r, dt), file=sys.stderr)
            prev = dt
            entry["d"] = dt.isoformat()
        elif cells[0].value:
            marker = str(cells[0].value).strip()
            if marker in DATE_LABELS:
                applied.add(marker)
                entry["dl"] = DATE_LABELS[marker]
            else:
                entry["m"] = marker

        cols = {}
        for i in range(1, 8):
            v = cells[i].value
            if v is None or not str(v).strip():
                continue
            items = apply_fixes(r, KEYS[i], split_items(str(v).strip()), applied)
            items = attach_archive(r, KEYS[i], items, applied)
            items = name_unmasked(KEYS[i], items)
            if cells[i].hyperlink:
                items[0]["l"] = cells[i].hyperlink.target
            cols[KEYS[i]] = items
        if cols:
            entry["c"] = cols
        entries.append(entry)

    expected = ({k[-1] for k in TITLE_FIXES} | {k[-1] for k in ROW_FIXES} |
                set(DATE_LABELS) | {k[-1] for k in EXTRA_LINKS})
    for miss in sorted(expected - applied):
        print("  note: correction no longer matches anything - %r" % miss, file=sys.stderr)
    print("applied %d/%d corrections" % (len(applied), len(expected)))

    print("season four: renumbered %d codes around the two-hour premiere"
          % season4.merge_entries(entries))

    moved, dropped = apply_reclass(entries)
    print("columns: moved %d releases out of Misc., dropped %d packed rows"
          % (moved, dropped))

    print("dates: moved %d items into the week two sources agree on"
          % apply_moves(entries))
    print("splits: %d extra entries pulled out of packed rows"
          % apply_splits(entries))
    print("bonus: %d novels read as Bonus rather than as a number"
          % apply_bonus(entries))

    extra = load_additions()
    if extra:
        print("additions: folded in %d releases the sheet has no row for"
              % apply_additions(entries, extra))

    print("eras: %d off-air weeks folded into the volume around them"
          % resolve_gaps(entries))

    return {"colLinks": col_links, "entries": entries}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--offline", action="store_true",
                    help="reuse the already-downloaded build/sheet.xlsx")
    args = ap.parse_args()

    if not args.offline:
        print("downloading sheet...")
        urllib.request.urlretrieve(XLSX_URL, XLSX)
    if not os.path.exists(XLSX):
        sys.exit("no build/sheet.xlsx - run without --offline first")

    data = build(XLSX)
    blob = json.dumps(data, ensure_ascii=False, separators=(",", ":"))

    html = open(HTML, encoding="utf-8").read()
    pat = re.compile(
        r'(<script type="application/json" id="timeline-data">).*?(</script>)', re.S)
    if not pat.search(html):
        sys.exit("could not find the timeline-data block in index.html")
    open(HTML, "w", encoding="utf-8").write(pat.sub(lambda m: m.group(1) + blob + m.group(2), html, count=1))

    items = sum(len(v) for e in data["entries"] for v in e.get("c", {}).values())
    print("wrote %d weeks / %d releases (%.1f KB) into index.html"
          % (len(data["entries"]), items, len(blob) / 1024))


if __name__ == "__main__":
    main()
